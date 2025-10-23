from nicegui import ui, app
import pandas as pd
import json, os, base64, requests, re
from io import BytesIO
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill

# === CONFIG ===
GITHUB_FILE = os.getenv("GITHUB_FILE", "keywords_memory.json")
GITHUB_REPO = os.getenv("GITHUB_REPO")
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")

# === RULES ===
RULES_A = {
    "LABORATORIO": ["analisi","emocromo","test","esame","coprolog","feci","giardia","leishmania","citolog","istolog","urinocolt","urine"],
    "VISITE": ["visita","controllo","consulto","dermatologic"],
    "FAR": ["meloxidyl","konclav","enrox","profenacarp","apoquel","osurnia","cylan","mometa","aristos","cytopoint","milbemax","stomorgyl","previcox","royal","stronghold","nexgard","procox"],
    "CHIRURGIA": ["intervento","chirurg","castraz","sterilizz","ovariect","detartrasi","estraz","biopsia","orchiettomia","odontostomat"],
    "DIAGNOSTICA PER IMMAGINI": ["rx","radiograf","eco","ecografia","tac"],
    "MEDICINA": ["terapia","terapie","flebo","day hospital","trattamento","emedog","cerenia","endovena","pressione"],
    "VACCINI": ["vacc","letifend","rabbia","trivalente","felv"],
    "CHIP": ["microchip","chip"],
    "ALTRE PRESTAZIONI": ["trasporto","eutanasia","unghie","cremazion","otoematoma","pet corner","ricette","medicazione","manualità"]
}

RULES_B = {
    "Visite domiciliari o presso allevamenti": ["visite domiciliari","allevamenti","domicilio"],
    "Visite ambulatoriali": ["visite ambulatoriali","terapia","trattamenti","vaccinazioni","ambulatorio","manualità","pet corner","visite","ricette","medicazione","microchip","controllo"],
    "Esami diagnostici per immagine": ["esami diagnostici per immagine","radiologia","eco","ecografia","tac","rx","raggi"],
    "Altri esami diagnostici": ["altri esami diagnostici","esami biochimici","laboratorio","malattie infettive","emocromo","prelievo"],
    "Interventi chirurgici": ["interventi chirurgici","avulsione","endoscopia","eutanasia","sedazione","anestesia","chirurgia","odontostomat","orchiettomia","asportazione","biopsia","ovariectomia"],
    "Assistenza al parto/ostetricia": ["assistenza al parto","ostetricia","parto"],
    "Attività di consulenza, perizia e collaborazione": ["attività di consulenza","perizia","collaborazione","telemedicina","consulto"],
    "Prestazioni di inseminazione artificiale": ["inseminazione artificiale"],
    "Altre attività": ["acconto"]
}

# === GITHUB ===
def github_load_json():
    try:
        if not (GITHUB_REPO and GITHUB_FILE):
            return {}
        url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}"
        headers = {"Authorization": f"token {GITHUB_TOKEN}"} if GITHUB_TOKEN else {}
        r = requests.get(url, headers=headers)
        if r.status_code == 200 and "content" in r.json():
            return json.loads(base64.b64decode(r.json()["content"]).decode("utf-8"))
    except Exception:
        pass
    return {}

def github_save_json(data: dict):
    try:
        if not (GITHUB_REPO and GITHUB_FILE and GITHUB_TOKEN):
            ui.notify("⚠️ GitHub non configurato.", color='warning')
            return
        url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}"
        headers = {"Authorization": f"token {GITHUB_TOKEN}"}
        get_res = requests.get(url, headers=headers)
        sha = get_res.json().get("sha") if get_res.status_code == 200 else None
        encoded = base64.b64encode(json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")).decode("utf-8")
        payload = {"message": "Aggiornamento dizionario Studio ISA", "content": encoded, "branch": "main"}
        if sha: payload["sha"] = sha
        requests.put(url, headers=headers, data=json.dumps(payload))
        ui.notify("✅ Dizionario aggiornato su GitHub!", color='green')
    except Exception as e:
        ui.notify(f"❌ Errore GitHub: {e}", color='red')

# === UTILS ===
def norm(s): return re.sub(r"\s+", " ", str(s).strip().lower())
def any_kw_in(t, kws): return any(k in t for k in kws)

def classify(desc, fam_val, mem, rules):
    if pd.notna(fam_val) and str(fam_val).strip():
        return str(fam_val).strip()
    d = norm(desc)
    for k,v in mem.items():
        if norm(k) in d:
            return v
    for cat,keys in rules.items():
        if any_kw_in(d, keys):
            return cat
    return list(rules.keys())[-1]

# === GLOBAL STATE ===
state = {"df": None, "ftype": None, "rules": None, "mem": github_load_json(), "updates": {}, "pending": [], "idx": 0}

# === FUNZIONI APP ===
def process_excel(e):
    file = e.content.read()
    df = pd.read_excel(BytesIO(file))
    state["df"] = df

    cols = [c.lower().strip() for c in df.columns]
    if any("prestazione" in c for c in cols) and any("totaleimpon" in c for c in cols):
        state["ftype"] = "B"
        state["rules"] = RULES_B
        ui.notify("🔍 Tipo file: VetsGo (B)")
    else:
        state["ftype"] = "A"
        state["rules"] = RULES_A
        ui.notify("🔍 Tipo file: DrVeto (A)")
    prepare_learning()

def prepare_learning():
    df = state["df"]
    mem = state["mem"]
    rules = state["rules"]
    ftype = state["ftype"]

    if ftype == "A":
        col_desc = next(c for c in df.columns if "descrizione" in c.lower())
        col_fam = next(c for c in df.columns if "famiglia" in c.lower())
        df["Categoria"] = df.apply(lambda r: classify(r[col_desc], r[col_fam], mem|state["updates"], rules), axis=1)
        base_col = col_desc
    else:
        col_prest = next(c for c in df.columns if "prestazioneprodotto" in c.replace(" ", "").lower())
        col_cat = next(c for c in df.columns if "categoria" in c.lower())
        df["Categoria"] = df.apply(lambda r: classify(r[col_prest], r[col_cat], mem|state["updates"], rules), axis=1)
        base_col = col_prest

    all_terms = sorted({str(v).strip() for v in df[base_col].dropna().unique()}, key=lambda s: s.casefold())
    pending = [t for t in all_terms if not any(norm(k) in norm(t) for k in (mem|state["updates"]).keys())]
    state["pending"] = pending
    show_learning()

def show_learning():
    content.clear()
    pending = state["pending"]
    idx = state["idx"]

    if not pending:
        ui.notify("✅ Tutti i termini classificati.")
        show_report()
        return

    term = pending[idx]
    ui.label(f"🧠 Nuovo termine ({idx+1}/{len(pending)}): {term}").classes("text-lg font-semibold mt-2")
    select_cat = ui.select(state["rules"].keys(), label="Seleziona categoria", value=list(state["rules"].keys())[0])

    with ui.row():
        ui.button("✅ Salva locale e prossimo", on_click=lambda: save_term(term, select_cat.value))
        ui.button("⏭️ Salta", on_click=next_term)
        ui.button("💾 Salva su GitHub", on_click=save_to_github)

def save_term(term, cat):
    state["updates"][term] = cat
    state["idx"] += 1
    ui.notify(f"💾 Salvato {term} → {cat}", color='green')
    show_learning()

def next_term():
    state["idx"] += 1
    show_learning()

def save_to_github():
    state["mem"].update(state["updates"])
    github_save_json(state["mem"])
    state["updates"] = {}
    ui.notify("✅ Dizionario aggiornato su GitHub!", color='green')

def show_report():
    df = state["df"]
    ftype = state["ftype"]
    content.clear()

    if ftype == "A":
        col_netto = next(c for c in df.columns if "netto" in c.lower() and "dopo" in c.lower())
        col_perc = next(c for c in df.columns if c.strip() == "%")
        studio = df.groupby("Categoria", dropna=False).agg({col_perc:"sum", col_netto:"sum"}).reset_index()
        studio.columns = ["FamigliaCategoria","Qtà","Netto"]
        tot_q, tot_n = studio["Qtà"].sum(), studio["Netto"].sum()
        studio["% Qtà"] = (studio["Qtà"]/tot_q*100).round(2)
        studio["% Netto"] = (studio["Netto"]/tot_n*100).round(2)
        studio = pd.concat([studio, pd.DataFrame([["Totale",tot_q,tot_n,100,100]], columns=studio.columns)], ignore_index=True)
        graph_title = "Somma Netto per FamigliaCategoria"
    else:
        col_imp = next(c for c in df.columns if "totaleimpon" in c.lower())
        col_iva = next(c for c in df.columns if "totaleconiva" in c.replace(" ", "").lower())
        col_tot = next(c for c in df.columns if "totale" in c.lower())
        studio = df.groupby("Categoria", dropna=False).agg({col_imp:"sum", col_iva:"sum", col_tot:"sum"}).reset_index()
        studio.columns = ["Categoria","TotaleImponibile","TotaleConIVA","Totale"]
        tot_t = studio["Totale"].sum()
        studio["% Totale"] = (studio["Totale"]/tot_t*100).round(2)
        studio = pd.concat([studio, pd.DataFrame([["Totale", studio["TotaleImponibile"].sum(), studio["TotaleConIVA"].sum(), tot_t, 100]], columns=studio.columns)], ignore_index=True)
        graph_title = "Somma Totale per Categoria"

    ui.table.from_pandas(studio).classes("mt-4")
    fig, ax = plt.subplots(figsize=(8,5))
    ax.bar(studio.iloc[:-1,0], studio.iloc[:-1,-2], color="skyblue")
    ax.set_title(graph_title)
    plt.xticks(rotation=45, ha="right")
    buf = BytesIO(); plt.tight_layout(); plt.savefig(buf, format="png"); buf.seek(0)
    ui.image(buf.getvalue())

    wb = Workbook(); ws = wb.active; ws.title = "Report"
    total_fill = PatternFill(start_color="FFF4B084", end_color="FFF4B084", fill_type="solid")
    start_row, start_col = 3, 2
    for j,h in enumerate(studio.columns,start=start_col):
        ws.cell(row=start_row,column=j,value=h).font = Font(bold=True)
    for i,row in enumerate(dataframe_to_rows(studio,index=False,header=False),start=start_row+1):
        for j,v in enumerate(row,start=start_col):
            ws.cell(row=i,column=j,value=v)
    tot_row_idx = start_row+len(studio)
    for j in range(start_col, start_col+len(studio.columns)):
        c=ws.cell(row=tot_row_idx,column=j); c.font=Font(bold=True); c.fill=total_fill
    img=XLImage(buf); img.anchor=f"A{tot_row_idx+3}"; ws.add_image(img)
    out=BytesIO(); wb.save(out)
    ui.download(data=out.getvalue(), filename=f"StudioISA_{ftype}_{datetime.now().year}.xlsx", label="⬇️ Scarica Excel")

# === UI ===
ui.page_title("Studio ISA – NiceGUI Edition")
ui.label("📊 Studio ISA – NiceGUI").classes("text-2xl font-bold mb-4")
upload = ui.upload(label="📁 Carica file Excel", multiple=False)
content = ui.column().classes("mt-6")

upload.on_upload(process_excel)

ui.run(title="Studio ISA – NiceGUI", reload=False, port=8080)
